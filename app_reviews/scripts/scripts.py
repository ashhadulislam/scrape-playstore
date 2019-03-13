from . import constants
from . import preprocess 
from .preprocess import PreProcess
import time
import pickle
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import pandas as pd
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.options import Options  
import os


def browser_functions(url):

    env=constants.env
    print("The env is ",env)



    chromedriver_loc = constants.chrome_driver_location
    if env=="dev":
        print("chromedriver ",constants.chrome_driver_location)
        chrome_options = Options()  
        chrome_options.add_argument("--headless")  
    elif env=="heroku_uat":
        print("chromedriver ",constants.chrome_driver_location)
        chrome_options = Options()
        chrome_options.binary_location = constants.google_chrome_bin
        chrome_options.add_argument("start-maximized"); # open Browser in maximized mode
        chrome_options.add_argument("disable-infobars"); # disabling infobars
        chrome_options.add_argument("--disable-extensions"); # disabling extensions
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage"); # overcome limited resource problems
        # chrome_options.add_argument('--disable-gpu')
    browser = webdriver.Chrome(executable_path=chromedriver_loc, chrome_options=chrome_options)        
    browser.get(url)


    no_of_pagedowns = constants.no_of_pagedowns
    elem = browser.find_element_by_tag_name("body")
    while no_of_pagedowns:
        print(no_of_pagedowns)
        dur=2
        elem.send_keys(Keys.END)
        time.sleep(dur)
        no_of_pagedowns-=1
        #click on the show more button if its there
        try:
            print("finding path")
            show_more=browser.find_element_by_xpath('//content[@class="CwaK9"]')
            print("find by span")
            show_more=show_more.find_element_by_tag_name("span")
            if show_more.text == "SHOW MORE":
                print("going to click")
                browser.execute_script("arguments[0].click();", show_more)
                print("just clicked")
    #             show_more.click()
        except Exception as e: 
            #code will come here when there are no buttons such as show more
            # print(e)
    #         print("Something wrong")
            continue

    source=browser.page_source
    print("got source successfully")
    browser.close()

    return source

def parse_html_page(source):

    soup = BeautifulSoup(source, 'html.parser')
    post_name_blocks=soup.find_all('div', class_='xKpxId zc7KVe')
    print("Number of posts = ",len(post_name_blocks))


    names=[]
    star_texts=[]
    star_numbers=[]
    review_dates=[]
    helpful_counts=[]
    count=0
    for post_name_block in post_name_blocks:

        #name of reviewer
        poster_name=post_name_block.find('span',class_='X43Kjb').text
        names.append(poster_name)

        #number of stars
        rating=post_name_block.find('div',class_='pf5lIe').findAll('div')
    #     print("rating",rating)
        rating_text=rating[0]
        rating_text=rating_text.get_attribute_list("aria-label")[0]
        rating_number=rating_text[6]

        star_texts.append(rating_text)
        star_numbers.append(rating_number)


        #get date of rating given
        rating_date=post_name_block.find('span',class_ = "p2TkOb").text    
        review_dates.append(rating_date)

        helpful_count=post_name_block.find("div",class_="jUL89d y92BAb").text
        helpful_counts.append(helpful_count)
        count+=1
        if count%100 == 0:
            print(count)

    review_texts=[]
    #for reviews

    reviews=soup.select('div.UD7Dzf > span')
    print("For this post")
    count=0
    while count<len(reviews):
        short_review=reviews[count].get_text(strip=True)
        count+=1
        if count==len(reviews):
            break
        long_review=reviews[count].get_text(strip=True)
        if len(short_review)!=0:
            review=short_review
        if len(long_review)!=0:
            review=long_review
        count+=1
        review_texts.append(review)
        if count%100 == 0:
            print(count)

    print("Final Review",len(review_texts))

    #very bad coding here
    min_length=min(len(names),len(review_dates),len(star_texts),len(star_numbers),len(helpful_counts),len(review_texts))
    print("The minimum length is ",min_length)


    d={}
    d["reviewer_name"]=names[0:min_length-1]
    d["review_date"]=review_dates[0:min_length-1]
    d["review_star_text"]=star_texts[0:min_length-1]
    d["review_star_count"]=star_numbers[0:min_length-1]
    d["review_helpful_count"]=helpful_counts[0:min_length-1]
    d["review"]=review_texts[0:min_length-1]


    for key, value in d.items():
        print(len(value))

    

    df = pd.DataFrame(data=d)
    print(df.shape)
    print(df.head())
    print("Got df successfully")
    return df




def get_reviews(url,stop_words):


    #adding the "get all review" part in the URL
    adder=constants._add_for_all_reviews
    if adder not in url:
        url=url+adder
    print("URL before call is",url)

    
    #get html source data after clicking on more review and tapping END
    html_source=browser_functions(url)
    if html_source==None:
        return False, None

    
    #parse the source to get the different components of the review
    df=parse_html_page(html_source)
    

    if df is None:
        return False, None

    
    df.to_csv(constants.output_location+"app_revs.csv")

    #here to do NLP
    column_name="review"
    print("going for text preprocessing")
    word_cloud_image_location=NLP_Flow(df,column_name,stop_words)
    print("word_cloud_image_location",word_cloud_image_location)



    #hold on, we will also write it to an excel

    #add the sheet to a list of sheets
    result_df_list=[]
    result_df_list.append(df)

    xls_name=os.path.join(constants.output_location, "result.xlsx")
    
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # # create Workbook object
    wb=Workbook()
    print("loaded some workbook")
    sheetnames=["result"]

    for n,dataf in enumerate(result_df_list):
        print("n is ",n)
        print(dataf.head())
        ws=wb.create_sheet(sheetnames[n])
        for r in dataframe_to_rows(dataf, index=True, header=True):
            ws.append(r)

    print("going to save at",xls_name)

    #remove the default sheet called "sheet"
    wb.remove(wb.get_sheet_by_name('Sheet'))
    wb.save(xls_name)

    print("written to xls")

    return True, xls_name,word_cloud_image_location




import numpy as np
from wordcloud import WordCloud
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt


def NLP_Flow(df,column_name,stop_words):
    pre_processor = PreProcess(df, column_name=column_name)
    data_check = pre_processor.clean_html()
    data_check = pre_processor.remove_non_ascii()
    data_check = pre_processor.remove_spaces()
    data_check = pre_processor.remove_punctuation()
    # data_check = pre_processor.stemming()
    # data_check = pre_processor.lemmatization()
    data_check = pre_processor.stop_words()

    print("After removing stop words")
    # print(data_check[column_name])

    #now remove stop words given manually
    print('stop_words',stop_words)

    if stop_words != None:
        pat = '\\b(?:{})\\b'.format('|'.join(stop_words))
        
        print('pat',pat)
        data_check['new'] = data_check[column_name].str.replace(pat, '')
        print('after removing manual stop words')
        print(data_check.head(60))

        data_check = data_check.rename(columns={column_name: 'old'+column_name})
        data_check = data_check.rename(columns={'new': column_name})


    print("column name is ",column_name)
    print(data_check.head())

    #get the tf idf matrix
    tfidf_transformer, tf_idf_matrix=preprocess.tf_idf(data_check,column_name="review")
    rows=tf_idf_matrix.todense().shape[0]
    columns=tf_idf_matrix.todense().shape[1]
    print(rows,columns)

    #convert the tf idf matrix to dataframe
    a = np.matrix(tf_idf_matrix.todense())
    df_with_tf_idf_value=pd.DataFrame(a)
    df_with_tf_idf_value.columns = tfidf_transformer.get_feature_names()

    #add each column to get total weights
    weight_word=df_with_tf_idf_value.sum()
    #convert to dictionary
    weight_word=weight_word.to_dict()
    for w in sorted(weight_word, key=weight_word.get, reverse=True):
        weight_word[w]=int(1000*weight_word[w])

    #Creating word cloud
    wordcloud = WordCloud(width=400,height=200, max_words=100,relative_scaling=1,normalize_plurals=False).generate_from_frequencies(weight_word)

    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis("off")
    print("current directory is ",os.getcwd())
    print(constants.output_location, " is there? ", os.path.exists(constants.output_location))
    print("going to save at ",constants.output_location)
    plot_location=os.path.join(constants.output_location,"wordcloud.png")
    plot_location_out=os.path.join(constants.output_location,"wordcloud_out.png")

    print("source location is ",os.getcwd())


    plt.savefig(plot_location)
    print("Saved at ",plot_location)

    with open(plot_location, 'rb') as f:
        data = f.read()

    with open(plot_location_out, 'wb') as f:
        f.write(data)




    print(plot_location , " is there? ", os.path.exists(plot_location))
    print(plot_location_out , " is there? ", os.path.exists(plot_location_out))
    return plot_location_out




